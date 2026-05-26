"""
baseline_config.py  (v3 — JSON-first)

เปลี่ยนหลักๆ:
  - เพิ่ม load_checks() / get_checks_for_version()
    อ่าน check definitions จาก /baselines/generated/*.json
    แทนการอ่าน Excel ตอน runtime
  - auto_detect_baseline() และ scan_data_directory() ยังอยู่
    แต่ใช้เฉพาะใน generate_baseline_json.py เท่านั้น
  - runtime ไม่แตะ .xlsx อีกเลย
"""

from __future__ import annotations

import json
import os
import re
from dataclasses import dataclass, field
from functools import lru_cache
from pathlib import Path
from typing import Optional

import openpyxl

# ---------------------------------------------------------------------------
# Constants (ใช้ใน auto_detect_baseline สำหรับ generate script เท่านั้น)
# ---------------------------------------------------------------------------

SHEET_TYPE_MAP: dict[str, str] = {
    "computer":           "computer",
    "user":               "user",
    "security template":  "security_template",
    "advanced audit":     "advanced_audit",
    "advanced auditing":  "advanced_audit",
    "firewall":           "firewall",
    "services":           "services",
    "applocker":          "skip",
    "applocker for dcs":  "applocker_dc",
    "information":        "skip",
    "revision history":   "skip",
}

TARGET_COLUMN_PRIORITY: list[str] = [
    "Member Server",
    "Domain Controller",
    "Windows 11 25H2",
    "Windows 11 24H2",
    "Windows 11",
    "Policy Value",
    "Windows 11 23H2",
    "Windows 10",
    "Windows Server 2022",
    "Windows Server 2019",
]

REG_INFO_CANDIDATES:    list[str] = ["Registry Information", "Reg Info", "Registry"]
POLICY_NAME_CANDIDATES: list[str] = ["Policy Setting Name", "Name", "Policy Name"]
POLICY_PATH_CANDIDATES: list[str] = ["Policy Path", "Path"]
SERVICE_NAME_CANDIDATES: list[str] = ["Name", "Service Name"]
SERVICE_TYPE_CANDIDATES: list[str] = ["Type", "Service Type"]

_FILENAME_RE = re.compile(
    r"MS[_ ]Security[_ ]Baseline[_ ](.+?)(?:\s*v[\d.]+)?\s*\.xlsx$",
    re.IGNORECASE,
)


# ---------------------------------------------------------------------------
# Data classes (ใช้ใน generate script + backward compat)
# ---------------------------------------------------------------------------

@dataclass
class SheetConfig:
    sheet_type:       str
    target_columns:   list[str]
    policy_name_col:  str = "Policy Setting Name"
    policy_path_col:  str = "Policy Path"
    reg_info_col:     str = "Registry Information"
    service_name_col: str = "Name"
    service_type_col: str = "Type"


@dataclass
class BaselineConfig:
    version_id:   str
    display_name: str
    filename:     str
    os_family:    str
    sheets:       dict[str, SheetConfig]
    skip_sheets:  set[str] = field(default_factory=lambda: {"Information", "Revision History"})


# ---------------------------------------------------------------------------
# auto_detect_baseline  (ใช้เฉพาะ generate script)
# ---------------------------------------------------------------------------

def _detect_os_family(filename: str) -> str:
    return "windows_server" if "server" in filename.lower() else "windows_client"


def _derive_version_id(filename: str) -> str:
    name = os.path.splitext(filename)[0]
    for p in ["MS_Security_Baseline_", "MS Security Baseline ", "MSSecurityBaseline"]:
        if name.lower().startswith(p.lower()):
            name = name[len(p):]
            break
    # ลบ suffix "MS Security Baseline" ที่อาจติดมาท้ายชื่อ
    name = re.sub(r'\s*\(?\bMS[\s_]Security[\s_]Baseline\b\)?', '', name, flags=re.IGNORECASE)
    return name.replace("_", " ").strip()


def _first_match(candidates: list[str], headers: list[str]) -> Optional[str]:
    for c in candidates:
        if c in headers:
            return c
    return None


def auto_detect_baseline(filepath: str) -> BaselineConfig:
    """อ่าน Excel → BaselineConfig  (ใช้เฉพาะ generate_baseline_json.py)"""
    filename   = os.path.basename(filepath)
    version_id = _derive_version_id(filename)
    os_family  = _detect_os_family(filename)

    wb     = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    sheets: dict[str, SheetConfig] = {}

    for sheet_name in wb.sheetnames:
        sheet_type = SHEET_TYPE_MAP.get(sheet_name.lower(), "skip")
        if sheet_type == "skip":
            continue

        ws   = wb[sheet_name]
        rows = list(ws.iter_rows(max_row=1, values_only=True))
        if not rows or not rows[0]:
            continue
        headers = [str(h).strip() if h is not None else "" for h in rows[0]]

        target_cols = [c for c in TARGET_COLUMN_PRIORITY if c in headers]
        if not target_cols:
            continue

        if sheet_type == "applocker_dc":
            sheets[sheet_name] = SheetConfig(
                sheet_type       = sheet_type,
                target_columns   = ["Domain Controller"],
                policy_name_col  = "Policy Setting Name",
                policy_path_col  = "Policy Path",
                reg_info_col     = "Policy Group or Registry Key",
                service_name_col = "Policy Setting",
                service_type_col = "Policy Type",
            )
            continue

        sheets[sheet_name] = SheetConfig(
            sheet_type       = sheet_type,
            target_columns   = target_cols,
            policy_name_col  = _first_match(POLICY_NAME_CANDIDATES, headers) or "Policy Setting Name",
            policy_path_col  = _first_match(POLICY_PATH_CANDIDATES, headers) or "Policy Path",
            reg_info_col     = _first_match(REG_INFO_CANDIDATES,    headers) or "Registry Information",
            service_name_col = _first_match(SERVICE_NAME_CANDIDATES, headers) or "Name",
            service_type_col = _first_match(SERVICE_TYPE_CANDIDATES, headers) or "Type",
        )

    wb.close()
    return BaselineConfig(
        version_id   = version_id,
        display_name = f"{version_id} (MS Security Baseline)",
        filename     = filename,
        os_family    = os_family,
        sheets       = sheets,
        skip_sheets  = {s for s in wb.sheetnames if SHEET_TYPE_MAP.get(s.lower(), "skip") == "skip"},
    )


def scan_data_directory(data_path: str) -> dict[str, BaselineConfig]:
    """Scan folder หา .xlsx  (ใช้เฉพาะ generate script)"""
    configs: dict[str, BaselineConfig] = {}
    if not os.path.isdir(data_path):
        return configs
    for fname in os.listdir(data_path):
        if not fname.lower().endswith(".xlsx"):
            continue
        if "security baseline" not in fname.lower() and "securitybaseline" not in fname.lower():
            continue
        fpath = os.path.join(data_path, fname)
        try:
            cfg = auto_detect_baseline(fpath)
            configs[cfg.version_id] = cfg
        except Exception as e:
            print(f"[baseline_config] skip {fname}: {e}")
    return configs


# ---------------------------------------------------------------------------
# resolve_target_col  (ยังใช้ได้ใน generate script)
# ---------------------------------------------------------------------------

def resolve_target_col(sheet_cfg: SheetConfig, columns: list,
                        role: str = "Member Server") -> Optional[str]:
    if role in columns:
        return role
    for candidate in sheet_cfg.target_columns:
        if candidate in columns:
            return candidate
    return None


# ===========================================================================
# JSON-first API  (ใช้ใน runtime / scanner)
# ===========================================================================

def _slug(value: str) -> str:
    return re.sub(r"[^a-zA-Z0-9]+", "-", value).strip("-").lower()


# ---------------------------------------------------------------------------
# load_checks — PUBLIC API สำหรับ scanner
# ---------------------------------------------------------------------------

# default path (override ได้ด้วย BASELINES_DIR env)
_DEFAULT_BASELINES_DIR = str(
    Path(__file__).resolve().parents[4] / "baselines" / "generated"
)

# cache: baseline_id → list[check]
_checks_cache: dict[str, list[dict]] = {}
_baselines_dir: str = ""


def _get_baselines_dir() -> str:
    return os.environ.get("BASELINES_DIR", _DEFAULT_BASELINES_DIR)


@lru_cache(maxsize=16)
def _load_json(filepath: str) -> dict:
    with open(filepath, "r", encoding="utf-8") as f:
        return json.load(f)


def load_checks(version_id: str, role: str = "Member Server") -> list[dict]:
    baselines_dir = _get_baselines_dir()

    # ลอง slug ตรงๆ ก่อน
    slug = _slug(version_id)
    filepath = os.path.join(baselines_dir, f"{slug}.json")

    # ถ้าไม่เจอ scan ทุกไฟล์แล้ว match
    if not os.path.exists(filepath):
        version_lower = version_id.strip().lower()
        for fname in sorted(os.listdir(baselines_dir)):
            if not fname.endswith(".json"):
                continue
            fpath = os.path.join(baselines_dir, fname)
            try:
                data = _load_json(fpath)
                bname = data.get("baseline_name", "").strip().lower()
                bid   = data.get("baseline_id",   "").strip().lower()
                # match ถ้าตรงกัน หรือ version_id เป็น substring ของ baseline_name
                if (bname == version_lower
                        or bid == slug
                        or version_lower in bname
                        or bname in version_lower):
                    filepath = fpath
                    break
            except Exception:
                continue

    if not os.path.exists(filepath):
        raise FileNotFoundError(
            f"ไม่พบ baseline definition: {filepath}\n"
            f"กรุณารัน: python scripts/generate_baseline_json.py"
        )

    data    = _load_json(filepath)
    checks  = data.get("checks", [])
    filtered = [
        c for c in checks
        if not c.get("applies_to") or _role_matches(c["applies_to"], role)
    ]
    return filtered


def _role_matches(applies_to: list[str], role: str) -> bool:
    """
    ตรวจว่า check นี้ apply กับ role ที่ต้องการหรือไม่
    role = "Member Server" → match กับ applies_to ที่มี "Member Server" หรือ "Windows 11..."
    role = "Domain Controller" → match กับ "Domain Controller"
    """
    role_lower = role.lower()
    for col in applies_to:
        col_lower = col.lower()
        if role_lower in col_lower:
            return True
        # Member Server ยังรองรับ Windows 11 / Windows 10
        if role_lower == "member server" and (
            "windows 11" in col_lower or "windows 10" in col_lower
            or "policy value" in col_lower
        ):
            return True
    return False


def list_available_versions(baselines_dir: str = "") -> list[dict]:
    """
    คืน list ของ baseline versions ที่มี JSON พร้อมใช้

    Returns
    -------
    list[dict]:  version_id, display_name, os_family, filename, check_count
    """
    bdir = baselines_dir or _get_baselines_dir()
    if not os.path.isdir(bdir):
        return []

    versions = []
    for fname in sorted(os.listdir(bdir)):
        if not fname.endswith(".json"):
            continue
        fpath = os.path.join(bdir, fname)
        try:
            data = _load_json(fpath)
            versions.append({
                "version_id":   data.get("baseline_name", fname),
                "display_name": data.get("baseline_name", fname),
                "os_family":    data.get("os_family", "unknown"),
                "filename":     fname,
                "check_count":  len(data.get("checks", [])),
            })
        except Exception as e:
            print(f"[baseline_config] skip {fname}: {e}")

    return versions


# ---------------------------------------------------------------------------
# Backward-compat shims  (main.py ยังเรียก load_configs / list_versions ได้)
# ---------------------------------------------------------------------------

_config_cache: dict[str, BaselineConfig] = {}
_cache_data_path: str = ""


def load_configs(data_path: str, force_reload: bool = False) -> dict[str, BaselineConfig]:
    """
    [compat] สแกน data_path หา .xlsx แล้ว return BaselineConfig
    ยังใช้งานได้แต่ runtime ไม่ควรเรียกแล้ว
    """
    global _config_cache, _cache_data_path
    if force_reload or _cache_data_path != data_path or not _config_cache:
        _config_cache    = scan_data_directory(data_path)
        _cache_data_path = data_path
    return _config_cache


def get_config(version_id: str, data_path: str = "") -> BaselineConfig:
    """[compat] ดึง BaselineConfig ตาม version_id จาก .xlsx"""
    configs = load_configs(data_path) if data_path else _config_cache
    cfg = configs.get(version_id)
    if cfg is None:
        raise ValueError(
            f"ไม่รองรับ version '{version_id}' "
            f"รองรับเฉพาะ: {list(configs.keys())}"
        )
    return cfg


def list_versions(data_path: str = "") -> list[dict]:
    """[compat] คืน list versions จาก .xlsx"""
    configs = load_configs(data_path) if data_path else _config_cache
    return [
        {
            "version_id":   cfg.version_id,
            "display_name": cfg.display_name,
            "filename":     cfg.filename,
            "os_family":    cfg.os_family,
        }
        for cfg in configs.values()
    ]