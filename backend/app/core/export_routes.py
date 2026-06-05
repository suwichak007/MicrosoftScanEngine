"""
export_routes.py
----------------
FastAPI router สำหรับ export ผลสแกนเป็น PDF หรือ CSV

Endpoints:
  GET /api/scan/history/{scan_id}/export/pdf
  GET /api/scan/history/{scan_id}/export/csv

ติดตั้ง dependency:
  pip install reportlab
"""

import csv
import io
import datetime
import re

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.core.database import SessionLocal
from app.core.security import get_current_user
from app.models.scan import ScanResult
from app.models.user import User

# reportlab imports
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT

router = APIRouter(prefix="/api/scan/history", tags=["export"])


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


def _get_scan_or_404(scan_id: int, current_user: User, db: Session) -> ScanResult:
    """ดึง scan result และตรวจสิทธิ์"""
    query = db.query(ScanResult).filter(ScanResult.id == scan_id)
    if current_user.role != "admin":
        query = query.filter(ScanResult.user_id == current_user.id)
    scan = query.first()
    if not scan:
        raise HTTPException(status_code=404, detail="ไม่พบผลการสแกนหรือไม่มีสิทธิ์เข้าถึง")
    return scan


def _classify_severity(key: str) -> str:
    """จัดระดับความรุนแรงจาก policy key (เหมือน frontend)"""
    lower = key.lower()
    critical_kw = ["remote desktop", "lsa protection", "credential", "ntlm", "kerberos", "bitlocker"]
    high_kw = [
        "network access", "network security", "user rights", "privilege",
        "logon", "encryption", "tls", "ssl", "rdp", "rpc",
        "anonymous", "guest", "sam", "domain member", "impersonate",
        "user account control", "restrict", "audit", "signing",
        "inactivity", "force shutdown",
    ]
    medium_kw = [
        "autoplay", "autorun", "internet explorer", "smartscreen", "activex",
        "printer", "bluetooth", "wifi", "hotspot", "ink workspace", "xbox",
        "cortana", "spotlight", "toast", "netbios", "icmp", "multicast",
    ]
    if any(k in lower for k in critical_kw):
        return "Critical"
    if lower.startswith("[advanced audit]"):
        return "Medium"
    if lower.startswith("[services]"):
        return "Low"
    if any(k in lower for k in high_kw):
        return "High"
    if any(k in lower for k in medium_kw):
        return "Medium"
    return "Low"


def _parse_details(details: dict) -> dict:
    """
    แยก details ออกเป็น pass_list, fail_list, manual_list
    พร้อม parse target/actual จาก string
    """
    pass_list   = []
    fail_list   = []
    manual_list = []

    for key, value in (details or {}).items():
        val_str = str(value)
        is_dict = isinstance(value, dict)
        section_match = key[1:key.index("]")] if key.startswith("[") and "]" in key else "General"
        name = key[key.index("]") + 1:].strip() if key.startswith("[") and "]" in key else key

        if is_dict:
            section_match = value.get("category") or value.get("section") or section_match
            name = value.get("check_name") or value.get("name") or name

        entry = {
            "key":      value.get("check_id") or key if is_dict else key,
            "name":     name,
            "section":  section_match,
            "severity": (
                str(value.get("severity") or "").title()
                if is_dict and value.get("severity")
                else _classify_severity(f"{section_match} {name} {key}")
            ),
            "status":   value.get("status", "") if is_dict else val_str,
            "target":   "",
            "actual":   "",
            "raw":      val_str,
        }

        if is_dict:
            entry["target"] = (
                value.get("expected_value")
                or value.get("target")
                or value.get("required_value")
                or ""
            )
            entry["actual"] = (
                value.get("current_value")
                or value.get("actual")
                or value.get("actual_value")
                or ""
            )
            entry["raw"] = value.get("raw_result") or val_str
            if not entry["actual"] and str(entry["status"]).strip().lower().startswith("pass") and entry["target"]:
                entry["actual"] = entry["target"]
        else:
            tgt = re.search(r"Target:\s*([^,)]+?)(?:\s*,|\s*\)|$)", val_str)
            act = re.search(r"Actual:\s*(.+?)(?:\s*\)\s*$|\s*$)", val_str)
            if tgt:
                entry["target"] = tgt.group(1).strip()
            if act:
                entry["actual"] = act.group(1).strip().rstrip(")")

        status_text = str(entry["status"] or "").strip()
        status_lower = status_text.lower()
        if status_lower.startswith("pass"):
            pass_list.append(entry)
        elif "manual" in status_lower or "not found" in status_lower:
            manual_list.append(entry)
        else:
            fail_list.append(entry)

    # เรียงตาม severity
    sev_order = {"Critical": 0, "High": 1, "Medium": 2, "Low": 3}
    fail_list.sort(key=lambda x: sev_order.get(x["severity"], 9))

    return {"pass": pass_list, "fail": fail_list, "manual": manual_list}


# ---------------------------------------------------------------------------
# PDF Generator
# ---------------------------------------------------------------------------

# สี theme (เข้ากับ UI)
COLOR_PRIMARY   = colors.HexColor("#c8813a")   # amber — brand color
COLOR_CRITICAL  = colors.HexColor("#dc2626")   # red
COLOR_HIGH      = colors.HexColor("#ea580c")   # orange
COLOR_MEDIUM    = colors.HexColor("#d97706")   # yellow
COLOR_LOW       = colors.HexColor("#2563eb")   # blue
COLOR_PASS      = colors.HexColor("#16a34a")   # green
COLOR_BG_HEADER = colors.HexColor("#1a1a2e")   # dark header
COLOR_BG_LIGHT  = colors.HexColor("#f8f9fa")   # light row
COLOR_TEXT      = colors.HexColor("#1f2937")   # body text

SEV_COLORS = {
    "Critical": COLOR_CRITICAL,
    "High":     COLOR_HIGH,
    "Medium":   COLOR_MEDIUM,
    "Low":      COLOR_LOW,
}


def _build_pdf(scan: ScanResult) -> bytes:
    """สร้าง PDF report จาก ScanResult และส่งคืนเป็น bytes"""
    buffer = io.BytesIO()
    page_w, page_h = A4
    margin = 20 * mm

    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=margin,
        rightMargin=margin,
        topMargin=margin,
        bottomMargin=margin,
        title=f"Security Report — {scan.target_name}",
        author="SecureScan",
    )

    styles = getSampleStyleSheet()

    # ─── Custom Styles ───────────────────────────────────────────────
    style_title = ParagraphStyle(
        "ReportTitle",
        parent=styles["Title"],
        fontSize=22,
        textColor=colors.white,
        alignment=TA_CENTER,
        spaceAfter=4,
    )
    style_subtitle = ParagraphStyle(
        "ReportSubtitle",
        parent=styles["Normal"],
        fontSize=10,
        textColor=colors.HexColor("#d1d5db"),
        alignment=TA_CENTER,
        spaceAfter=0,
    )
    style_section = ParagraphStyle(
        "SectionHead",
        parent=styles["Heading2"],
        fontSize=12,
        textColor=COLOR_PRIMARY,
        spaceBefore=10,
        spaceAfter=4,
        borderPad=4,
    )
    style_body = ParagraphStyle(
        "Body",
        parent=styles["Normal"],
        fontSize=8.5,
        textColor=COLOR_TEXT,
        leading=13,
    )
    style_small = ParagraphStyle(
        "Small",
        parent=styles["Normal"],
        fontSize=7.5,
        textColor=colors.HexColor("#6b7280"),
        leading=11,
    )
    style_cell = ParagraphStyle(
        "Cell",
        parent=styles["Normal"],
        fontSize=8,
        textColor=COLOR_TEXT,
        leading=11,
        wordWrap="CJK",
    )

    story = []
    parsed   = _parse_details(scan.details)
    details  = scan.details or {}
    total    = len(details)
    n_pass   = len(parsed["pass"])
    n_fail   = len(parsed["fail"])
    n_manual = len(parsed["manual"])
    score    = scan.score or 0
    scan_date = scan.scan_date.strftime("%d/%m/%Y %H:%M") if scan.scan_date else "—"

    # ─── Header block ────────────────────────────────────────────────
    header_data = [[
        Paragraph("🛡 Security Baseline Scan Report", style_title),
    ]]
    header_table = Table(header_data, colWidths=[page_w - 2 * margin])
    header_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), COLOR_BG_HEADER),
        ("TOPPADDING",    (0, 0), (-1, -1), 16),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 16),
        ("LEFTPADDING",   (0, 0), (-1, -1), 12),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 12),
        ("ROUNDEDCORNERS", [4]),
    ]))
    story.append(header_table)
    story.append(Spacer(1, 6))

    # subtitle row
    sub_text = (
        f"Target: {scan.target_name or '—'}  ·  "
        f"Version: {scan.version or '—'}  ·  "
        f"Hostname: {scan.hostname or '—'}  ·  "
        f"Date: {scan_date}"
    )
    story.append(Paragraph(sub_text, style_small))
    story.append(Spacer(1, 10))

    # ─── Score + Summary cards ───────────────────────────────────────
    score_color = (
        COLOR_PASS if score >= 70
        else COLOR_MEDIUM if score >= 40
        else COLOR_CRITICAL
    )

    summary_data = [
        [
            Paragraph(f'<font size="28" color="{score_color.hexval()}"><b>{score}%</b></font><br/><font size="8" color="#6b7280">Health Score</font>', style_cell),
            Paragraph(f'<font size="20" color="{COLOR_PASS.hexval()}"><b>{n_pass}</b></font><br/><font size="8" color="#6b7280">Pass</font>', style_cell),
            Paragraph(f'<font size="20" color="{COLOR_CRITICAL.hexval()}"><b>{n_fail}</b></font><br/><font size="8" color="#6b7280">Fail</font>', style_cell),
            Paragraph(f'<font size="20" color="{COLOR_MEDIUM.hexval()}"><b>{n_manual}</b></font><br/><font size="8" color="#6b7280">Manual</font>', style_cell),
            Paragraph(f'<font size="20" color="{COLOR_TEXT.hexval()}"><b>{total}</b></font><br/><font size="8" color="#6b7280">Total</font>', style_cell),
        ]
    ]
    col_w = (page_w - 2 * margin) / 5
    summary_table = Table(summary_data, colWidths=[col_w] * 5)
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), COLOR_BG_LIGHT),
        ("ALIGN",      (0, 0), (-1, -1), "CENTER"),
        ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",    (0, 0), (-1, -1), 12),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
        ("ROUNDEDCORNERS", [4]),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 6))

    # severity breakdown
    sev_counts = {"Critical": 0, "High": 0, "Medium": 0, "Low": 0}
    for item in parsed["fail"]:
        sev_counts[item["severity"]] = sev_counts.get(item["severity"], 0) + 1

    sev_data = [[
        Paragraph(
            f'<font color="{SEV_COLORS[s].hexval()}"><b>{s}: {sev_counts[s]}</b></font>',
            style_cell
        )
        for s in ["Critical", "High", "Medium", "Low"]
    ]]
    sev_table = Table(sev_data, colWidths=[(page_w - 2 * margin) / 4] * 4)
    sev_table.setStyle(TableStyle([
        ("ALIGN",      (0, 0), (-1, -1), "CENTER"),
        ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",    (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("BACKGROUND", (0, 0), (-1, -1), colors.white),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
    ]))
    story.append(sev_table)
    story.append(Spacer(1, 12))
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#e5e7eb")))
    story.append(Spacer(1, 8))

    # ─── Failed Checks Table ─────────────────────────────────────────
    if parsed["fail"]:
        story.append(Paragraph("Failed Checks", style_section))
        story.append(Spacer(1, 4))

        col_widths = [
            18 * mm,   # Severity
            55 * mm,   # Policy Name
            25 * mm,   # Section
            30 * mm,   # Current Value
            30 * mm,   # Required
        ]

        # header row
        header_row = [
            Paragraph("<b>Severity</b>",      style_cell),
            Paragraph("<b>Policy</b>",         style_cell),
            Paragraph("<b>Section</b>",        style_cell),
            Paragraph("<b>Current Value</b>",  style_cell),
            Paragraph("<b>Required</b>",       style_cell),
        ]
        table_data = [header_row]

        for i, item in enumerate(parsed["fail"]):
            sev_color  = SEV_COLORS.get(item["severity"], COLOR_LOW)
            row_bg     = COLOR_BG_LIGHT if i % 2 == 0 else colors.white

            sev_para   = Paragraph(
                f'<font color="{sev_color.hexval()}"><b>{item["severity"]}</b></font>',
                style_cell
            )
            # ตัดชื่อถ้ายาวเกิน
            name_text  = item["name"][:80] + "…" if len(item["name"]) > 80 else item["name"]
            actual_text = item["actual"][:40] + "…" if len(item["actual"]) > 40 else item["actual"] or "Not Configured"
            target_text = item["target"][:40] + "…" if len(item["target"]) > 40 else item["target"] or "—"

            table_data.append([
                sev_para,
                Paragraph(name_text,    style_cell),
                Paragraph(item["section"], style_cell),
                Paragraph(actual_text,  style_cell),
                Paragraph(target_text,  style_cell),
            ])

        fail_table = Table(table_data, colWidths=col_widths, repeatRows=1)

        # build row background styles
        row_styles = [
            ("BACKGROUND", (0, 0), (-1, 0), COLOR_BG_HEADER),
            ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
            ("FONTSIZE",   (0, 0), (-1, -1), 8),
            ("TOPPADDING",    (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("LEFTPADDING",   (0, 0), (-1, -1), 5),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 5),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#e5e7eb")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]
        for i in range(1, len(table_data)):
            if i % 2 == 0:
                row_styles.append(("BACKGROUND", (0, i), (-1, i), COLOR_BG_LIGHT))

        fail_table.setStyle(TableStyle(row_styles))
        story.append(fail_table)
        story.append(Spacer(1, 12))

    # ─── Manual Check Table (collapsed) ──────────────────────────────
    if parsed["manual"]:
        story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#e5e7eb")))
        story.append(Spacer(1, 8))
        story.append(Paragraph("Manual Check Required", style_section))
        story.append(Spacer(1, 4))

        manual_col_w = [(page_w - 2 * margin) * r for r in [0.55, 0.25, 0.20]]
        manual_data  = [[
            Paragraph("<b>Policy</b>",   style_cell),
            Paragraph("<b>Section</b>",  style_cell),
            Paragraph("<b>Status</b>",   style_cell),
        ]]
        for item in parsed["manual"]:
            name_text = item["name"][:90] + "…" if len(item["name"]) > 90 else item["name"]
            manual_data.append([
                Paragraph(name_text,       style_cell),
                Paragraph(item["section"], style_cell),
                Paragraph(
                    f'<font color="{COLOR_MEDIUM.hexval()}">Manual Check</font>',
                    style_cell
                ),
            ])

        manual_table = Table(manual_data, colWidths=manual_col_w, repeatRows=1)
        manual_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#374151")),
            ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
            ("FONTSIZE",   (0, 0), (-1, -1), 8),
            ("TOPPADDING",    (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING",   (0, 0), (-1, -1), 5),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 5),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#e5e7eb")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, COLOR_BG_LIGHT]),
        ]))
        story.append(manual_table)
        story.append(Spacer(1, 12))

    # ─── Footer ──────────────────────────────────────────────────────
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#e5e7eb")))
    story.append(Spacer(1, 6))
    footer_text = (
        f"Generated by SecureScan  ·  "
        f"{datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}  ·  "
        f"Microsoft Security Baseline"
    )
    story.append(Paragraph(footer_text, style_small))

    doc.build(story)
    buffer.seek(0)
    return buffer.read()


# ---------------------------------------------------------------------------
# CSV Generator
# ---------------------------------------------------------------------------

def _build_csv(scan: ScanResult) -> bytes:
    """สร้าง CSV report จาก ScanResult ส่งคืนเป็น bytes (UTF-8 BOM สำหรับ Excel)"""
    output  = io.StringIO()
    writer  = csv.writer(output)

    scan_date = scan.scan_date.strftime("%d/%m/%Y %H:%M") if scan.scan_date else ""

    # metadata header
    writer.writerow(["Security Baseline Scan Report"])
    writer.writerow(["Target",   scan.target_name or ""])
    writer.writerow(["Hostname", scan.hostname    or ""])
    writer.writerow(["Version",  scan.version     or ""])
    writer.writerow(["Score",    f"{scan.score}%"])
    writer.writerow(["Date",     scan_date])
    writer.writerow([])

    # column header
    writer.writerow([
        "Policy Key",
        "Policy Name",
        "Section",
        "Severity",
        "Status",
        "Current Value",
        "Required Value",
        "Raw Result",
    ])

    parsed = _parse_details(scan.details)
    all_items = parsed["fail"] + parsed["manual"] + parsed["pass"]

    for item in all_items:
        writer.writerow([
            item["key"],
            item["name"],
            item["section"],
            item["severity"],
            "Fail"   if item in parsed["fail"]
            else "Manual" if item in parsed["manual"]
            else "Pass",
            item["actual"],
            item["target"],
            item["raw"],
        ])

    # UTF-8 BOM ทำให้ Excel เปิดได้ถูกต้อง
    return ("\ufeff" + output.getvalue()).encode("utf-8")


def _build_xlsx(scan: ScanResult) -> bytes:
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_results = wb.create_sheet("Results")

    scan_date = scan.scan_date.strftime("%d/%m/%Y %H:%M") if scan.scan_date else ""
    parsed = _parse_details(scan.details)
    all_items = parsed["fail"] + parsed["manual"] + parsed["pass"]

    title_fill = PatternFill("solid", fgColor="1A1A2E")
    header_fill = PatternFill("solid", fgColor="E8D9C8")
    pass_fill = PatternFill("solid", fgColor="DDEEDC")
    fail_fill = PatternFill("solid", fgColor="F8D8D4")
    manual_fill = PatternFill("solid", fgColor="FCE8C2")
    white_font = Font(color="FFFFFF", bold=True)
    header_font = Font(color="1F2937", bold=True)

    ws_summary["A1"] = "Security Baseline Scan Report"
    ws_summary["A1"].fill = title_fill
    ws_summary["A1"].font = white_font
    ws_summary.merge_cells("A1:B1")
    summary_rows = [
        ("Target", scan.target_name or ""),
        ("Hostname", scan.hostname or ""),
        ("Version", scan.version or ""),
        ("Score", f"{scan.score or 0}%"),
        ("Date", scan_date),
        ("Total Checks", len(scan.details or {})),
        ("Pass", len(parsed["pass"])),
        ("Fail", len(parsed["fail"])),
        ("Manual", len(parsed["manual"])),
    ]
    for idx, (label, value) in enumerate(summary_rows, start=3):
        ws_summary.cell(row=idx, column=1, value=label).font = header_font
        ws_summary.cell(row=idx, column=2, value=value)
    ws_summary.column_dimensions["A"].width = 22
    ws_summary.column_dimensions["B"].width = 60

    headers = [
        "Policy Key",
        "Policy Name",
        "Section",
        "Severity",
        "Status",
        "Current Value",
        "Required Value",
        "Raw Result",
    ]
    ws_results.append(headers)
    for cell in ws_results[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for item in all_items:
        status = "Fail" if item in parsed["fail"] else "Manual" if item in parsed["manual"] else "Pass"
        ws_results.append([
            item["key"],
            item["name"],
            item["section"],
            item["severity"],
            status,
            item["actual"],
            item["target"],
            item["raw"],
        ])
        fill = pass_fill if status == "Pass" else manual_fill if status == "Manual" else fail_fill
        for cell in ws_results[ws_results.max_row]:
            cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws_results.freeze_panes = "A2"
    ws_results.auto_filter.ref = ws_results.dimensions
    widths = [42, 58, 26, 14, 12, 28, 28, 70]
    for idx, width in enumerate(widths, start=1):
        ws_results.column_dimensions[get_column_letter(idx)].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------

@router.get("/{scan_id}/export/pdf")
async def export_pdf(
    scan_id:      int,
    db:           Session = Depends(get_db),
    current_user: User    = Depends(get_current_user),
):
    """Export ผลการสแกนเป็น PDF"""
    scan = _get_scan_or_404(scan_id, current_user, db)

    try:
        pdf_bytes = _build_pdf(scan)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"ไม่สามารถสร้าง PDF ได้: {e}")

    filename  = f"scan-report-{scan_id}-{datetime.date.today()}.pdf"
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{scan_id}/export/csv")
async def export_csv(
    scan_id:      int,
    db:           Session = Depends(get_db),
    current_user: User    = Depends(get_current_user),
):
    """Export ผลการสแกนเป็น CSV"""
    scan = _get_scan_or_404(scan_id, current_user, db)

    try:
        csv_bytes = _build_csv(scan)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"ไม่สามารถสร้าง CSV ได้: {e}")

    filename = f"scan-report-{scan_id}-{datetime.date.today()}.csv"
    return StreamingResponse(
        io.BytesIO(csv_bytes),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{scan_id}/export/xlsx")
async def export_xlsx(
    scan_id:      int,
    db:           Session = Depends(get_db),
    current_user: User    = Depends(get_current_user),
):
    scan = _get_scan_or_404(scan_id, current_user, db)

    try:
        xlsx_bytes = _build_xlsx(scan)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Cannot create XLSX: {e}")

    filename = f"scan-report-{scan_id}-{datetime.date.today()}.xlsx"
    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

